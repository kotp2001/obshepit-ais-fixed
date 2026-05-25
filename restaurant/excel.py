import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from .models import Order
from datetime import datetime, timedelta
from collections import defaultdict

HEADER_FILL  = PatternFill(start_color="1a1a4a", end_color="1a1a4a", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFD700")
CENTER       = Alignment(horizontal='center', vertical='center')

def _parse_period(request):
    """Возвращает (start_date, end_date, label) по GET-параметрам запроса."""
    period = request.GET.get('period', 'week')
    today  = datetime.now().date()

    if period == 'day':
        return today, today, 'сегодня'
    elif period == 'month':
        return today - timedelta(days=30), today, 'месяц'
    elif period == 'custom':
        df = request.GET.get('date_from')
        dt = request.GET.get('date_to')
        start = datetime.strptime(df, '%Y-%m-%d').date() if df else today - timedelta(days=7)
        end   = datetime.strptime(dt, '%Y-%m-%d').date() if dt else today
        return start, end, f'{start.strftime("%d.%m.%Y")}–{end.strftime("%d.%m.%Y")}'
    else:  # week
        return today - timedelta(days=7), today, 'неделя'

def _style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

def export_orders_excel(request):
    start_date, end_date, label = _parse_period(request)

    orders = Order.objects.filter(
        status='paid',
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).select_related('table').prefetch_related('items__dish').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    _style_header(ws, ['ID', 'Стол', 'Дата', 'Время', 'Сумма (₽)', 'Оплата', 'Позиции'])

    for row, order in enumerate(orders, 2):
        items_text = ", ".join(f"{i.dish.name} x{i.quantity}" for i in order.items.all())
        ws.cell(row=row, column=1, value=order.id)
        ws.cell(row=row, column=2, value=order.table.number)
        ws.cell(row=row, column=3, value=order.created_at.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=4, value=order.created_at.strftime('%H:%M'))
        ws.cell(row=row, column=5, value=float(order.total_amount))
        ws.cell(row=row, column=6, value=dict(Order.PAYMENT_CHOICES).get(order.payment_method, ''))
        ws.cell(row=row, column=7, value=items_text)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    fname = f'orders_{start_date}_{end_date}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={fname}'
    wb.save(response)
    return response

def export_popular_excel(request):
    start_date, end_date, label = _parse_period(request)

    orders = Order.objects.filter(
        status='paid',
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).prefetch_related('items__dish')

    dish_count = defaultdict(int)
    dish_revenue = defaultdict(float)
    for order in orders:
        for item in order.items.all():
            dish_count[item.dish.name]   += item.quantity
            dish_revenue[item.dish.name] += float(item.price * item.quantity)

    sorted_dishes = sorted(dish_count.items(), key=lambda x: -x[1])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Популярные блюда"

    _style_header(ws, ['Место', 'Блюдо', 'Продано (шт.)', 'Выручка (₽)'])

    for idx, (name, count) in enumerate(sorted_dishes, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=name)
        ws.cell(row=idx+1, column=3, value=count)
        ws.cell(row=idx+1, column=4, value=round(dish_revenue[name], 2))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    fname = f'popular_{start_date}_{end_date}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={fname}'
    wb.save(response)
    return response
